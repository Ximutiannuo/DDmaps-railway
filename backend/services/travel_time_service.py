"""
行驶时间数据库服务模块（Excel 解析和导出）
"""
import json
import io
from datetime import datetime
from typing import Dict, Any, List, Optional
from uuid import uuid4

try:
    import openpyxl  # type: ignore
except ImportError:
    openpyxl = None

from backend.utils.logger import logger


def _require_openpyxl() -> None:
    """确保已安装 openpyxl"""
    if openpyxl is None:
        raise RuntimeError('当前环境未安装 openpyxl，请先运行 pip install openpyxl')


TRAVEL_TIME_NUMERIC_FIELDS = {
    'duration_minutes',
    'distance_m',
    'average_speed_kmph',
    'custom_speed_kmph',
    'speed_setting_kmph',
    'estimated_minutes',
    'estimated_delta_minutes',
    'route_efficiency_score',
    'weight_tons',
    'width_m',
    'path_edge_count',
    'path_node_count'
}

TRAVEL_TIME_DATETIME_FIELDS = {
    'start_time',
    'arrival_time',
    'route_request_time',
    'created_at'
}

TRAVEL_TIME_VALID_FIELDS = {
    'record_id',
    'driver_id',
    'driver_name',
    'vehicle_id',
    'vehicle_type',
    'start_node',
    'target_node',
    'route_key',
    'duration_minutes',
    'distance_m',
    'average_speed_kmph',
    'custom_speed_kmph',
    'speed_setting_kmph',
    'custom_speed_source',
    'estimated_minutes',
    'estimated_delta_minutes',
    'start_time',
    'arrival_time',
    'route_request_time',
    'route_efficiency_score',
    'path_edge_count',
    'path_node_count',
    'weight_tons',
    'width_m',
    'source',
    'weekday',
    'hour_of_day',
    'created_at',
    'path_nodes',
    'path_edges',
    'data_version'
}

TRAVEL_TIME_HEADER_ALIASES = {
    'driver_id': 'driver_id',
    '司机id': 'driver_id',
    '司机编号': 'driver_id',
    'driver': 'driver_name',
    '司机姓名': 'driver_name',
    '司机': 'driver_name',
    'driver_name': 'driver_name',
    'vehicle_id': 'vehicle_id',
    '车辆id': 'vehicle_id',
    '车辆编号': 'vehicle_id',
    '车辆类型': 'vehicle_type',
    'vehicle_type': 'vehicle_type',
    'start_node': 'start_node',
    '起点': 'start_node',
    'target_node': 'target_node',
    '终点': 'target_node',
    'route_key': 'route_key',
    '路线': 'route_key',
    'duration_minutes': 'duration_minutes',
    'duration': 'duration_minutes',
    '耗时': 'duration_minutes',
    '耗时(分钟)': 'duration_minutes',
    'distance_m': 'distance_m',
    'distance': 'distance_m',
    '距离': 'distance_m',
    '距离(m)': 'distance_m',
    'average_speed_kmph': 'average_speed_kmph',
    'avg_speed': 'average_speed_kmph',
    '平均速度': 'average_speed_kmph',
    '平均速度(km/h)': 'average_speed_kmph',
    'custom_speed_kmph': 'custom_speed_kmph',
    '自定义速度': 'custom_speed_kmph',
    '速度设定': 'speed_setting_kmph',
    'speed_setting_kmph': 'speed_setting_kmph',
    'custom_speed_source': 'custom_speed_source',
    '速度来源': 'custom_speed_source',
    'estimated_minutes': 'estimated_minutes',
    'estimated_delta_minutes': 'estimated_delta_minutes',
    'start_time': 'start_time',
    '出发时间': 'start_time',
    'arrival_time': 'arrival_time',
    '到达时间': 'arrival_time',
    'route_request_time': 'route_request_time',
    'route_efficiency_score': 'route_efficiency_score',
    'path_edge_count': 'path_edge_count',
    'path_node_count': 'path_node_count',
    'path_nodes': 'path_nodes',
    '路径节点': 'path_nodes',
    'path_edges': 'path_edges',
    '路径道路': 'path_edges',
    'weight_tons': 'weight_tons',
    '载重': 'weight_tons',
    'width_m': 'width_m',
    '宽度': 'width_m',
    'source': 'source',
    'weekday': 'weekday',
    'hour_of_day': 'hour_of_day',
    'created_at': 'created_at',
    '记录时间': 'created_at',
    'record_id': 'record_id',
    '记录id': 'record_id',
    'data_version': 'data_version'
}


def calculate_average_speed_kmph(distance_m: Optional[float], duration_minutes: Optional[float]) -> Optional[float]:
    """根据距离和耗时计算平均速度（km/h）"""
    if not distance_m or distance_m <= 0:
        return None
    if not duration_minutes or duration_minutes <= 0:
        return None
    hours = duration_minutes / 60.0
    if hours <= 0:
        return None
    speed = (distance_m / 1000.0) / hours
    return round(speed, 2)


def normalize_travel_time_records(raw_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """对导入的行驶时间记录进行字段规范化"""
    normalized_records: List[Dict[str, Any]] = []
    for rec in raw_records:
        if not isinstance(rec, dict):
            continue
        data = dict(rec)

        for field in TRAVEL_TIME_DATETIME_FIELDS:
            if field in data and data[field] is not None:
                value = data[field]
                if isinstance(value, datetime):
                    data[field] = value.isoformat()
                else:
                    data[field] = str(value).strip()

        for field in TRAVEL_TIME_NUMERIC_FIELDS:
            if field in data and data[field] not in (None, ''):
                try:
                    data[field] = float(data[field])
                except (TypeError, ValueError):
                    data[field] = None

        if data.get('distance_m') is not None:
            data['distance_m'] = float(round(float(data['distance_m']), 3))
        if data.get('duration_minutes') is not None:
            data['duration_minutes'] = float(round(float(data['duration_minutes']), 3))
        if data.get('average_speed_kmph') is not None:
            data['average_speed_kmph'] = float(round(float(data['average_speed_kmph']), 3))

        if data.get('average_speed_kmph') in (None, '', 0):
            computed_speed = calculate_average_speed_kmph(
                data.get('distance_m'),
                data.get('duration_minutes')
            )
            if computed_speed is not None:
                data['average_speed_kmph'] = computed_speed

        # 处理 path_nodes：支持 JSON 格式和箭头分隔格式
        if isinstance(data.get('path_nodes'), str):
            nodes_str = data['path_nodes'].strip()
            if nodes_str:
                try:
                    # 先尝试解析 JSON 格式
                    parsed = json.loads(nodes_str)
                    if isinstance(parsed, list):
                        data['path_nodes'] = parsed
                    else:
                        # 如果不是列表，尝试用箭头分割
                        nodes_str = nodes_str.replace('→', '->')
                        data['path_nodes'] = [item.strip() for item in nodes_str.split('->') if item.strip()]
                except (json.JSONDecodeError, ValueError):
                    # JSON 解析失败，使用箭头分割格式
                    nodes_str = nodes_str.replace('→', '->')
                    data['path_nodes'] = [item.strip() for item in nodes_str.split('->') if item.strip()]
            else:
                data['path_nodes'] = []
        
        # 处理 path_edges：支持 JSON 格式和箭头分隔格式
        if isinstance(data.get('path_edges'), str):
            edges_str = data['path_edges'].strip()
            if edges_str:
                try:
                    # 先尝试解析 JSON 格式
                    parsed = json.loads(edges_str)
                    if isinstance(parsed, list):
                        data['path_edges'] = parsed
                    else:
                        # 如果不是列表，尝试用箭头分割
                        edges_str = edges_str.replace('→', '->')
                        data['path_edges'] = [item.strip() for item in edges_str.split('->') if item.strip()]
                except (json.JSONDecodeError, ValueError):
                    # JSON 解析失败，使用箭头分割格式
                    edges_str = edges_str.replace('→', '->')
                    data['path_edges'] = [item.strip() for item in edges_str.split('->') if item.strip()]
            else:
                data['path_edges'] = []

        if isinstance(data.get('weekday'), str):
            try:
                data['weekday'] = int(data['weekday'])
            except ValueError:
                data['weekday'] = None
        if isinstance(data.get('hour_of_day'), str):
            try:
                data['hour_of_day'] = int(data['hour_of_day'])
            except ValueError:
                data['hour_of_day'] = None

        if data.get('path_nodes') and isinstance(data['path_nodes'], list):
            data['path_node_count'] = len(data['path_nodes'])
        if data.get('path_edges') and isinstance(data['path_edges'], list):
            data['path_edge_count'] = len(data['path_edges'])

        if data.get('start_node') and data.get('target_node'):
            data.setdefault('route_key', f"{data['start_node']}->{data['target_node']}")

        if not data.get('record_id'):
            data['record_id'] = f'TR-IMPORT-{uuid4().hex}'
        data.setdefault('data_version', 1)
        cleaned_record: Dict[str, Any] = {}
        for key, value in data.items():
            if key in TRAVEL_TIME_VALID_FIELDS:
                cleaned_record[key] = value
        normalized_records.append(cleaned_record)
    return normalized_records


def parse_travel_time_excel(file_storage) -> List[Dict[str, Any]]:
    """从 Excel 中解析行驶时间记录"""
    _require_openpyxl()
    workbook = openpyxl.load_workbook(file_storage, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError('Excel 文件为空')

    header_row = rows[0]
    header_map = {}
    for idx, header in enumerate(header_row):
        if header is None:
            continue
        header_str = str(header).strip()
        if not header_str:
            continue
        normalized_key = TRAVEL_TIME_HEADER_ALIASES.get(header_str.lower()) or TRAVEL_TIME_HEADER_ALIASES.get(header_str)
        if not normalized_key and header_str in TRAVEL_TIME_VALID_FIELDS:
            normalized_key = header_str
        if not normalized_key and header_str.lower() in TRAVEL_TIME_VALID_FIELDS:
            normalized_key = header_str.lower()
        if normalized_key and normalized_key in TRAVEL_TIME_VALID_FIELDS:
            header_map[idx] = normalized_key

    if not header_map:
        raise ValueError('Excel 文件缺少有效的表头，请确认第一行包含字段名称')

    parsed_records: List[Dict[str, Any]] = []
    for row in rows[1:]:
        record: Dict[str, Any] = {}
        for idx, key in header_map.items():
            if idx >= len(row):
                continue
            value = row[idx]
            if value is None:
                continue
            if isinstance(value, datetime):
                record[key] = value.isoformat()
            elif isinstance(value, (list, tuple)):
                record[key] = list(value)
            else:
                record[key] = str(value).strip() if isinstance(value, str) else value

        if record:
            parsed_records.append(record)

    return parsed_records


def build_travel_time_excel(records: List[Dict[str, Any]]) -> io.BytesIO:
    """将行驶时间记录导出为 Excel 工作簿"""
    _require_openpyxl()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'TravelTime'

    headers = [
        'record_id',
        'driver_id',
        'driver_name',
        'vehicle_id',
        'vehicle_type',
        'start_node',
        'target_node',
        'route_key',
        'duration_minutes',
        'distance_m',
        'average_speed_kmph',
        'custom_speed_kmph',
        'speed_setting_kmph',
        'custom_speed_source',
        'estimated_minutes',
        'estimated_delta_minutes',
        'start_time',
        'arrival_time',
        'route_request_time',
        'route_efficiency_score',
        'path_edge_count',
        'path_node_count',
        'weight_tons',
        'width_m',
        'source',
        'weekday',
        'hour_of_day',
        'created_at',
        'path_nodes',
        'path_edges'
    ]
    worksheet.append(headers)

    def format_path(value):
        if value is None:
            return ''
        if isinstance(value, (list, dict)):
            try:
                return json.dumps(value, ensure_ascii=False)
            except (TypeError, ValueError):
                pass
        if isinstance(value, list):
            formatted = []
            for item in value:
                if isinstance(item, dict):
                    name = item.get('name') or item.get('id') or json.dumps(item, ensure_ascii=False)
                    formatted.append(str(name))
                else:
                    formatted.append(str(item))
            return ' -> '.join(formatted)
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    for record in records:
        row = []
        for key in headers:
            value = record.get(key)
            if key in ('path_nodes', 'path_edges'):
                value = format_path(value)
            row.append(value)
        worksheet.append(row)

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream

